"""
agents/workflow_agent/sheet_connector.py

Purpose:
    Reads/writes Google Sheets, Airtable, Excel/CSV rows, performs duplicate checks,
    and exports row data for William / Jarvis Workflow Agent.

Where to place:
    agents/workflow_agent/sheet_connector.py

Required dependencies:
    Core Python:
        - csv
        - json
        - logging
        - os
        - uuid
        - datetime
        - pathlib
        - dataclasses
        - typing

    Optional for Excel:
        pip install openpyxl

    Optional for Google Sheets:
        pip install gspread google-auth

    Optional for Airtable:
        pip install requests

How to test:
    python -m py_compile agents/workflow_agent/sheet_connector.py

    Minimal local Excel/CSV test:
        from agents.workflow_agent.sheet_connector import SheetConnector

        connector = SheetConnector()
        ctx = {"user_id": "user_1", "workspace_id": "workspace_1"}

        result = connector.write_rows(
            context=ctx,
            provider="excel",
            target={
                "file_path": "test_leads.xlsx",
                "sheet_name": "Leads"
            },
            rows=[
                {"name": "John Doe", "phone": "+15551234567", "service": "Web Design"}
            ],
            create_if_missing=True
        )
        print(result)

Agent/module completion:
    Agent/Module: Workflow Agent
    File Completed: sheet_connector.py
    Completion: 47.6%
    Completed Files: ['workflow_agent.py', 'n8n_connector.py', 'workflow_builder.py', 'trigger_engine.py', 'action_router.py', 'app_connector.py', 'webhook_manager.py', 'form_pipeline.py', 'crm_connector.py', 'sheet_connector.py']
    Remaining Files: ['whatsapp_connector.py', 'email_connector.py', 'notification_engine.py', 'condition_engine.py', 'scheduler.py', 'workflow_monitor.py', 'retry_handler.py', 'workflow_templates.py', 'workflow_memory.py', 'approval_gate.py', 'config.py']
    Next Recommended File: agents/workflow_agent/whatsapp_connector.py

FILE COMPLETE
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import tempfile
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple, Union


# =============================================================================
# Optional imports
# =============================================================================

try:
    import requests  # type: ignore
except Exception:  # pragma: no cover
    requests = None  # type: ignore

try:
    import openpyxl  # type: ignore
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

try:
    import gspread  # type: ignore
except Exception:  # pragma: no cover
    gspread = None  # type: ignore


# =============================================================================
# Safe BaseAgent fallback
# =============================================================================

try:
    from agents.base_agent import BaseAgent  # type: ignore
except Exception:  # pragma: no cover
    try:
        from agents.base import BaseAgent  # type: ignore
    except Exception:  # pragma: no cover

        class BaseAgent:  # type: ignore
            """
            Import-safe fallback BaseAgent.

            This fallback keeps this file safe to import before the complete William
            agent framework exists. When the real BaseAgent is available, it will
            be used automatically.
            """

            def __init__(self, *args: Any, **kwargs: Any) -> None:
                self.agent_name = kwargs.get("agent_name", self.__class__.__name__)
                self.agent_type = kwargs.get("agent_type", "workflow_agent")
                self.agent_id = kwargs.get("agent_id", self.agent_name.lower())


# =============================================================================
# Constants
# =============================================================================

SUPPORTED_PROVIDERS = {"google_sheets", "airtable", "excel", "csv"}
WRITE_OPERATIONS = {"write_row", "write_rows", "update_rows", "delete_rows", "upsert_rows"}
READ_OPERATIONS = {"read_rows", "find_duplicates", "export_rows"}
DEFAULT_MAX_ROWS = 5000
DEFAULT_TIMEOUT_SECONDS = 30
SAFE_FILENAME_PATTERN = re.compile(r"^[a-zA-Z0-9._/\-\\ ]+$")


# =============================================================================
# Data structures
# =============================================================================

@dataclass
class SheetConnectorConfig:
    """
    Configuration for SheetConnector.

    This config intentionally avoids hardcoded secrets. Credentials should be
    supplied via secure runtime configuration, environment variables, encrypted
    app connector config, or a Security Agent approved secret resolver.
    """

    default_timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS
    max_rows_per_read: int = DEFAULT_MAX_ROWS
    max_rows_per_write: int = 10000
    allow_local_file_access: bool = True
    export_base_dir: str = "storage/exports/workflow_agent"
    audit_enabled: bool = True
    events_enabled: bool = True
    require_security_for_external_write: bool = True
    require_security_for_local_write: bool = False
    require_security_for_exports: bool = False
    allowed_local_extensions: Tuple[str, ...] = (".xlsx", ".csv")
    redact_sensitive_values_in_logs: bool = True


@dataclass
class SheetTarget:
    """
    Target descriptor for a supported spreadsheet-like provider.

    Google Sheets target fields:
        spreadsheet_id: Required unless spreadsheet_name is used by caller extension.
        worksheet_name: Optional, defaults to first sheet.
        credentials_json: Optional dictionary.
        credentials_file: Optional path.
        service_account_env: Optional env var containing JSON credentials.

    Airtable target fields:
        base_id: Required.
        table_name: Required.
        api_key: Optional. Prefer secret manager.
        api_key_env: Optional env var name.
        view: Optional.
        base_url: Optional.

    Excel/CSV target fields:
        file_path: Required.
        sheet_name: Optional for Excel.
    """

    provider: str
    spreadsheet_id: Optional[str] = None
    spreadsheet_name: Optional[str] = None
    worksheet_name: Optional[str] = None

    base_id: Optional[str] = None
    table_name: Optional[str] = None
    view: Optional[str] = None
    base_url: str = "https://api.airtable.com/v0"

    file_path: Optional[str] = None
    sheet_name: Optional[str] = None

    credentials_json: Optional[Dict[str, Any]] = None
    credentials_file: Optional[str] = None
    service_account_env: Optional[str] = None

    api_key: Optional[str] = None
    api_key_env: Optional[str] = None

    extra: Dict[str, Any] = field(default_factory=dict)


@dataclass
class DuplicateCheckConfig:
    """
    Duplicate detection configuration.
    """

    key_fields: List[str]
    case_sensitive: bool = False
    trim_whitespace: bool = True
    normalize_phone: bool = True
    ignore_empty_keys: bool = True


@dataclass
class ExportConfig:
    """
    Export configuration for rows.
    """

    export_format: str = "csv"
    file_name: Optional[str] = None
    include_metadata: bool = True
    overwrite: bool = False


# =============================================================================
# Utility helpers
# =============================================================================

def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_list_of_dicts(rows: Any) -> List[Dict[str, Any]]:
    if rows is None:
        return []

    if isinstance(rows, Mapping):
        return [dict(rows)]

    if not isinstance(rows, list):
        raise ValueError("Rows must be a dictionary or a list of dictionaries.")

    normalized: List[Dict[str, Any]] = []
    for index, row in enumerate(rows):
        if not isinstance(row, Mapping):
            raise ValueError(f"Row at index {index} must be a dictionary.")
        normalized.append(dict(row))
    return normalized


def _flatten_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    try:
        return json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        return str(value)


def _normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\D+", "", str(value))


def _safe_mkdir(path: Union[str, Path]) -> None:
    Path(path).mkdir(parents=True, exist_ok=True)


def _safe_read_json_env(env_var: str) -> Optional[Dict[str, Any]]:
    raw = os.getenv(env_var)
    if not raw:
        return None
    try:
        loaded = json.loads(raw)
        if isinstance(loaded, dict):
            return loaded
    except Exception:
        return None
    return None


def _redact(value: Any) -> Any:
    if value is None:
        return None
    text = str(value)
    if len(text) <= 8:
        return "***"
    return f"{text[:3]}***{text[-3:]}"


def _safe_file_path(path_value: str) -> Path:
    """
    Basic path safety guard.

    This does not replace application-level storage isolation, but helps prevent
    obviously unsafe paths from being accidentally used.
    """

    if not path_value or not isinstance(path_value, str):
        raise ValueError("file_path is required.")

    if "\x00" in path_value:
        raise ValueError("Invalid file path.")

    if not SAFE_FILENAME_PATTERN.match(path_value):
        raise ValueError("File path contains unsupported characters.")

    return Path(path_value).expanduser().resolve()


def _dict_without_none(data: Dict[str, Any]) -> Dict[str, Any]:
    return {key: value for key, value in data.items() if value is not None}


# =============================================================================
# Main connector
# =============================================================================

class SheetConnector(BaseAgent):
    """
    Production-ready SheetConnector for William / Jarvis Workflow Agent.

    Responsibilities:
        - Read rows from Google Sheets, Airtable, Excel, and CSV.
        - Write/append rows to Google Sheets, Airtable, Excel, and CSV.
        - Update rows where supported.
        - Check duplicates using configurable fields.
        - Export rows to CSV, JSON, or Excel.
        - Preserve SaaS user/workspace isolation through context validation.
        - Prepare payloads for Security Agent, Verification Agent, Memory Agent,
          Dashboard/API, Agent Registry, and Master Agent routing.

    Important architecture notes:
        - This file does not hardcode secrets.
        - External write actions can require Security Agent approval.
        - Local file write/export permissions can be controlled by config.
        - Public methods return structured dictionaries:
            {
                "success": bool,
                "message": str,
                "data": dict,
                "error": dict | None,
                "metadata": dict
            }
    """

    def __init__(
        self,
        config: Optional[Union[SheetConnectorConfig, Dict[str, Any]]] = None,
        security_client: Optional[Any] = None,
        verification_client: Optional[Any] = None,
        memory_client: Optional[Any] = None,
        audit_logger: Optional[Callable[[Dict[str, Any]], None]] = None,
        event_emitter: Optional[Callable[[Dict[str, Any]], None]] = None,
        logger: Optional[logging.Logger] = None,
        **kwargs: Any,
    ) -> None:
        super().__init__(
            agent_name=kwargs.pop("agent_name", "SheetConnector"),
            agent_type=kwargs.pop("agent_type", "workflow_agent"),
            agent_id=kwargs.pop("agent_id", "workflow_agent.sheet_connector"),
            **kwargs,
        )

        if config is None:
            self.config = SheetConnectorConfig()
        elif isinstance(config, SheetConnectorConfig):
            self.config = config
        elif isinstance(config, dict):
            self.config = SheetConnectorConfig(**config)
        else:
            raise TypeError("config must be SheetConnectorConfig, dict, or None.")

        self.security_client = security_client
        self.verification_client = verification_client
        self.memory_client = memory_client
        self.audit_logger = audit_logger
        self.event_emitter = event_emitter
        self.logger = logger or logging.getLogger(__name__)

    # -------------------------------------------------------------------------
    # Required compatibility hooks
    # -------------------------------------------------------------------------

    def _validate_task_context(self, context: Mapping[str, Any]) -> Dict[str, Any]:
        """
        Validate SaaS context.

        Every task touching user/workspace data must carry user_id and
        workspace_id to prevent cross-tenant data mixing.
        """

        if not isinstance(context, Mapping):
            return self._error_result(
                message="Invalid task context.",
                code="invalid_context",
                details={"reason": "context must be a dictionary"},
            )

        user_id = context.get("user_id")
        workspace_id = context.get("workspace_id")

        if not user_id or not workspace_id:
            return self._error_result(
                message="Missing required SaaS isolation context.",
                code="missing_context",
                details={
                    "required": ["user_id", "workspace_id"],
                    "received": list(context.keys()),
                },
            )

        return self._safe_result(
            message="Task context validated.",
            data={
                "user_id": str(user_id),
                "workspace_id": str(workspace_id),
                "request_id": str(context.get("request_id") or uuid.uuid4()),
                "actor_id": context.get("actor_id") or user_id,
                "role": context.get("role"),
                "source": context.get("source", "workflow_agent"),
            },
        )

    def _requires_security_check(
        self,
        operation: str,
        provider: str,
        context: Mapping[str, Any],
        target: Optional[Mapping[str, Any]] = None,
    ) -> bool:
        """
        Decide whether Security Agent approval is required.

        External writes are sensitive because they can modify third-party data.
        Exports may be sensitive because they create files containing user data.
        """

        operation = str(operation or "").lower()
        provider = str(provider or "").lower()

        if context.get("security_approved") is True:
            return False

        if operation in WRITE_OPERATIONS:
            if provider in {"google_sheets", "airtable"}:
                return self.config.require_security_for_external_write
            if provider in {"excel", "csv"}:
                return self.config.require_security_for_local_write

        if operation in {"export_rows"}:
            return self.config.require_security_for_exports

        return False

    def _request_security_approval(
        self,
        context: Mapping[str, Any],
        operation: str,
        provider: str,
        target: Optional[Mapping[str, Any]] = None,
        payload_summary: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Request Security Agent approval.

        If no Security Agent client is attached, this method returns a blocked
        response for operations that require approval. For dry-run or already
        approved contexts, callers should skip this check.
        """

        approval_payload = {
            "event_type": "security_approval_request",
            "agent": "SheetConnector",
            "operation": operation,
            "provider": provider,
            "context": {
                "user_id": context.get("user_id"),
                "workspace_id": context.get("workspace_id"),
                "actor_id": context.get("actor_id"),
                "request_id": context.get("request_id"),
            },
            "target": self._sanitize_target_for_logs(target or {}),
            "payload_summary": dict(payload_summary or {}),
            "created_at": _utc_now_iso(),
        }

        if self.security_client is None:
            return self._error_result(
                message="Security approval is required before this sheet operation can run.",
                code="security_approval_required",
                details=approval_payload,
                metadata={"requires_security_approval": True},
            )

        try:
            if hasattr(self.security_client, "approve_action"):
                approval = self.security_client.approve_action(approval_payload)
            elif hasattr(self.security_client, "request_approval"):
                approval = self.security_client.request_approval(approval_payload)
            elif callable(self.security_client):
                approval = self.security_client(approval_payload)
            else:
                return self._error_result(
                    message="Invalid Security Agent client.",
                    code="invalid_security_client",
                    details={"approval_payload": approval_payload},
                    metadata={"requires_security_approval": True},
                )

            if isinstance(approval, Mapping) and approval.get("approved") is True:
                return self._safe_result(
                    message="Security approval granted.",
                    data={"approval": dict(approval)},
                    metadata={"requires_security_approval": False},
                )

            return self._error_result(
                message="Security approval denied or unavailable.",
                code="security_approval_denied",
                details={"approval": approval, "approval_payload": approval_payload},
                metadata={"requires_security_approval": True},
            )

        except Exception as exc:
            return self._error_result(
                message="Security approval request failed.",
                code="security_approval_failed",
                details={"exception": str(exc), "approval_payload": approval_payload},
                metadata={"requires_security_approval": True},
            )

    def _prepare_verification_payload(
        self,
        context: Mapping[str, Any],
        operation: str,
        provider: str,
        result: Mapping[str, Any],
        target: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Prepare Verification Agent payload.

        The Verification Agent can use this to confirm write counts, duplicate
        outcomes, export file existence, and downstream workflow correctness.
        """

        return {
            "verification_type": "workflow_sheet_operation",
            "agent": "SheetConnector",
            "operation": operation,
            "provider": provider,
            "success": bool(result.get("success")),
            "message": result.get("message"),
            "context": {
                "user_id": context.get("user_id"),
                "workspace_id": context.get("workspace_id"),
                "request_id": context.get("request_id"),
            },
            "target": self._sanitize_target_for_logs(target or {}),
            "data_summary": self._summarize_result_data(result.get("data", {})),
            "created_at": _utc_now_iso(),
        }

    def _prepare_memory_payload(
        self,
        context: Mapping[str, Any],
        operation: str,
        provider: str,
        result: Mapping[str, Any],
        target: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Prepare Memory Agent-compatible payload.

        This does not store raw spreadsheet rows by default. It stores useful
        workflow context only, reducing privacy risk.
        """

        data = result.get("data") if isinstance(result.get("data"), Mapping) else {}

        return {
            "memory_type": "workflow_sheet_context",
            "agent": "SheetConnector",
            "operation": operation,
            "provider": provider,
            "context": {
                "user_id": context.get("user_id"),
                "workspace_id": context.get("workspace_id"),
                "request_id": context.get("request_id"),
            },
            "summary": {
                "success": bool(result.get("success")),
                "message": result.get("message"),
                "row_count": data.get("row_count"),
                "written_count": data.get("written_count"),
                "updated_count": data.get("updated_count"),
                "duplicate_count": data.get("duplicate_count"),
                "export_path": data.get("export_path"),
            },
            "target": self._sanitize_target_for_logs(target or {}),
            "created_at": _utc_now_iso(),
        }

    def _emit_agent_event(self, event: Mapping[str, Any]) -> None:
        """
        Emit dashboard/API/registry friendly event.

        Safe no-op if no event emitter is attached.
        """

        if not self.config.events_enabled:
            return

        event_payload = dict(event)
        event_payload.setdefault("agent", "SheetConnector")
        event_payload.setdefault("created_at", _utc_now_iso())

        try:
            if self.event_emitter:
                self.event_emitter(event_payload)
            else:
                self.logger.debug("SheetConnector event: %s", event_payload)
        except Exception as exc:
            self.logger.warning("Failed to emit SheetConnector event: %s", exc)

    def _log_audit_event(self, event: Mapping[str, Any]) -> None:
        """
        Log audit event.

        This is where the Dashboard/API, audit log store, or Agent Registry can
        track spreadsheet access while preserving user/workspace boundaries.
        """

        if not self.config.audit_enabled:
            return

        audit_payload = dict(event)
        audit_payload.setdefault("agent", "SheetConnector")
        audit_payload.setdefault("created_at", _utc_now_iso())

        try:
            if self.audit_logger:
                self.audit_logger(audit_payload)
            else:
                self.logger.info("SheetConnector audit: %s", audit_payload)
        except Exception as exc:
            self.logger.warning("Failed to log SheetConnector audit event: %s", exc)

    def _safe_result(
        self,
        message: str,
        data: Optional[Mapping[str, Any]] = None,
        metadata: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Standard structured success result.
        """

        return {
            "success": True,
            "message": message,
            "data": dict(data or {}),
            "error": None,
            "metadata": {
                "agent": "SheetConnector",
                "timestamp": _utc_now_iso(),
                **dict(metadata or {}),
            },
        }

    def _error_result(
        self,
        message: str,
        code: str = "sheet_connector_error",
        details: Optional[Mapping[str, Any]] = None,
        metadata: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Standard structured error result.
        """

        return {
            "success": False,
            "message": message,
            "data": {},
            "error": {
                "code": code,
                "details": dict(details or {}),
            },
            "metadata": {
                "agent": "SheetConnector",
                "timestamp": _utc_now_iso(),
                **dict(metadata or {}),
            },
        }

    # -------------------------------------------------------------------------
    # Public metadata methods for Master Agent / Registry / Loader
    # -------------------------------------------------------------------------

    def get_capabilities(self) -> Dict[str, Any]:
        """
        Return capability manifest for Agent Registry and Master Agent routing.
        """

        return self._safe_result(
            message="SheetConnector capabilities loaded.",
            data={
                "agent": "SheetConnector",
                "module": "workflow_agent",
                "file": "sheet_connector.py",
                "providers": sorted(SUPPORTED_PROVIDERS),
                "operations": [
                    "read_rows",
                    "write_row",
                    "write_rows",
                    "update_rows",
                    "upsert_rows",
                    "find_duplicates",
                    "export_rows",
                    "normalize_rows",
                ],
                "supports_user_workspace_isolation": True,
                "supports_security_approval": True,
                "supports_verification_payload": True,
                "supports_memory_payload": True,
                "safe_to_import_without_optional_dependencies": True,
            },
        )

    def health_check(self) -> Dict[str, Any]:
        """
        Lightweight health check.
        """

        return self._safe_result(
            message="SheetConnector health check completed.",
            data={
                "status": "ok",
                "optional_dependencies": {
                    "requests": requests is not None,
                    "openpyxl": openpyxl is not None,
                    "gspread": gspread is not None,
                },
                "supported_providers": sorted(SUPPORTED_PROVIDERS),
            },
        )

    # -------------------------------------------------------------------------
    # Main public methods
    # -------------------------------------------------------------------------

    def read_rows(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: Mapping[str, Any],
        limit: Optional[int] = None,
        offset: int = 0,
        filters: Optional[Mapping[str, Any]] = None,
        include_empty: bool = False,
    ) -> Dict[str, Any]:
        """
        Read rows from Google Sheets, Airtable, Excel, or CSV.
        """

        operation = "read_rows"
        provider = self._normalize_provider(provider)

        context_result = self._validate_task_context(context)
        if not context_result["success"]:
            return context_result

        ctx = context_result["data"]
        target_obj_result = self._build_target(provider, target)
        if not target_obj_result["success"]:
            return target_obj_result

        target_obj: SheetTarget = target_obj_result["data"]["target_obj"]
        read_limit = self._safe_limit(limit)

        self._audit_start(ctx, operation, provider, target_obj)

        try:
            if provider == "google_sheets":
                rows = self._google_read_rows(target_obj)
            elif provider == "airtable":
                rows = self._airtable_read_rows(target_obj, limit=read_limit, offset=offset)
            elif provider == "excel":
                rows = self._excel_read_rows(target_obj)
            elif provider == "csv":
                rows = self._csv_read_rows(target_obj)
            else:
                return self._error_result(
                    message=f"Unsupported provider: {provider}",
                    code="unsupported_provider",
                    details={"provider": provider, "supported": sorted(SUPPORTED_PROVIDERS)},
                )

            rows = self._apply_filters(rows, filters or {})
            if not include_empty:
                rows = [row for row in rows if any(value not in ("", None) for value in row.values())]

            if offset and provider != "airtable":
                rows = rows[offset:]

            rows = rows[:read_limit]

            result = self._safe_result(
                message="Rows read successfully.",
                data={
                    "provider": provider,
                    "row_count": len(rows),
                    "rows": rows,
                    "offset": offset,
                    "limit": read_limit,
                },
                metadata={
                    "operation": operation,
                    "user_id": ctx["user_id"],
                    "workspace_id": ctx["workspace_id"],
                    "verification_payload": self._prepare_verification_payload(
                        ctx,
                        operation,
                        provider,
                        {"success": True, "message": "Rows read successfully.", "data": {"row_count": len(rows)}},
                        asdict(target_obj),
                    ),
                },
            )

            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

        except Exception as exc:
            result = self._error_result(
                message="Failed to read rows.",
                code="read_rows_failed",
                details={"exception": str(exc), "provider": provider},
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

    def write_row(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: Mapping[str, Any],
        row: Mapping[str, Any],
        create_if_missing: bool = True,
        dry_run: bool = False,
        dedupe: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Write a single row.
        """

        return self.write_rows(
            context=context,
            provider=provider,
            target=target,
            rows=[dict(row)],
            create_if_missing=create_if_missing,
            dry_run=dry_run,
            dedupe=dedupe,
        )

    def write_rows(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: Mapping[str, Any],
        rows: Sequence[Mapping[str, Any]],
        create_if_missing: bool = True,
        dry_run: bool = False,
        dedupe: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Append rows to supported provider.

        For SaaS safety:
            - user_id and workspace_id are required.
            - external writes can require Security Agent approval.
            - duplicate checks can be performed before write.
        """

        operation = "write_rows"
        provider = self._normalize_provider(provider)

        context_result = self._validate_task_context(context)
        if not context_result["success"]:
            return context_result
        ctx = context_result["data"]

        target_obj_result = self._build_target(provider, target)
        if not target_obj_result["success"]:
            return target_obj_result
        target_obj: SheetTarget = target_obj_result["data"]["target_obj"]

        try:
            normalized_rows = self.normalize_rows(rows)
        except Exception as exc:
            return self._error_result(
                message="Invalid rows.",
                code="invalid_rows",
                details={"exception": str(exc)},
            )

        if not normalized_rows:
            return self._error_result(
                message="No rows provided to write.",
                code="empty_rows",
            )

        if len(normalized_rows) > self.config.max_rows_per_write:
            return self._error_result(
                message="Too many rows for one write operation.",
                code="max_rows_per_write_exceeded",
                details={
                    "received": len(normalized_rows),
                    "max_rows_per_write": self.config.max_rows_per_write,
                },
            )

        if dedupe:
            dedupe_result = self._dedupe_before_write(
                ctx,
                provider,
                target_obj,
                normalized_rows,
                dedupe,
            )
            if not dedupe_result["success"]:
                return dedupe_result
            normalized_rows = dedupe_result["data"]["rows_to_write"]

        payload_summary = {
            "row_count": len(normalized_rows),
            "columns": self._collect_headers(normalized_rows),
            "dry_run": dry_run,
        }

        if self._requires_security_check(operation, provider, ctx, asdict(target_obj)):
            approval = self._request_security_approval(
                context=ctx,
                operation=operation,
                provider=provider,
                target=asdict(target_obj),
                payload_summary=payload_summary,
            )
            if not approval["success"]:
                return approval

        self._audit_start(ctx, operation, provider, target_obj, payload_summary=payload_summary)

        if dry_run:
            result = self._safe_result(
                message="Dry run completed. No rows were written.",
                data={
                    "provider": provider,
                    "written_count": 0,
                    "row_count": len(normalized_rows),
                    "rows_preview": normalized_rows[:5],
                    "dry_run": True,
                },
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

        try:
            if provider == "google_sheets":
                written_count = self._google_append_rows(target_obj, normalized_rows, create_if_missing)
            elif provider == "airtable":
                written_count = self._airtable_create_rows(target_obj, normalized_rows)
            elif provider == "excel":
                written_count = self._excel_append_rows(target_obj, normalized_rows, create_if_missing)
            elif provider == "csv":
                written_count = self._csv_append_rows(target_obj, normalized_rows, create_if_missing)
            else:
                return self._error_result(
                    message=f"Unsupported provider: {provider}",
                    code="unsupported_provider",
                    details={"provider": provider, "supported": sorted(SUPPORTED_PROVIDERS)},
                )

            result = self._safe_result(
                message="Rows written successfully.",
                data={
                    "provider": provider,
                    "written_count": written_count,
                    "row_count": len(normalized_rows),
                    "dry_run": False,
                },
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

        except Exception as exc:
            result = self._error_result(
                message="Failed to write rows.",
                code="write_rows_failed",
                details={"exception": str(exc), "provider": provider},
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

    def update_rows(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: Mapping[str, Any],
        match: Mapping[str, Any],
        updates: Mapping[str, Any],
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        """
        Update rows that match exact field values.

        Supported:
            - Airtable
            - Excel
            - CSV
            - Google Sheets

        Google Sheets update is implemented by reading all records and rewriting
        changed rows through worksheet.update.
        """

        operation = "update_rows"
        provider = self._normalize_provider(provider)

        context_result = self._validate_task_context(context)
        if not context_result["success"]:
            return context_result
        ctx = context_result["data"]

        target_obj_result = self._build_target(provider, target)
        if not target_obj_result["success"]:
            return target_obj_result
        target_obj: SheetTarget = target_obj_result["data"]["target_obj"]

        if not match:
            return self._error_result(
                message="Update match criteria is required.",
                code="missing_match_criteria",
            )

        if not updates:
            return self._error_result(
                message="Update values are required.",
                code="missing_updates",
            )

        payload_summary = {
            "match_fields": list(match.keys()),
            "update_fields": list(updates.keys()),
            "dry_run": dry_run,
        }

        if self._requires_security_check(operation, provider, ctx, asdict(target_obj)):
            approval = self._request_security_approval(
                context=ctx,
                operation=operation,
                provider=provider,
                target=asdict(target_obj),
                payload_summary=payload_summary,
            )
            if not approval["success"]:
                return approval

        self._audit_start(ctx, operation, provider, target_obj, payload_summary=payload_summary)

        try:
            if provider == "google_sheets":
                updated_count = self._google_update_rows(target_obj, match, updates, dry_run)
            elif provider == "airtable":
                updated_count = self._airtable_update_rows(target_obj, match, updates, dry_run)
            elif provider == "excel":
                updated_count = self._excel_update_rows(target_obj, match, updates, dry_run)
            elif provider == "csv":
                updated_count = self._csv_update_rows(target_obj, match, updates, dry_run)
            else:
                return self._error_result(
                    message=f"Unsupported provider: {provider}",
                    code="unsupported_provider",
                    details={"provider": provider},
                )

            result = self._safe_result(
                message="Rows updated successfully." if not dry_run else "Dry run completed. No rows were updated.",
                data={
                    "provider": provider,
                    "updated_count": updated_count if not dry_run else 0,
                    "matched_count": updated_count,
                    "dry_run": dry_run,
                },
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

        except Exception as exc:
            result = self._error_result(
                message="Failed to update rows.",
                code="update_rows_failed",
                details={"exception": str(exc), "provider": provider},
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider, target_obj, result)
            return result

    def upsert_rows(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: Mapping[str, Any],
        rows: Sequence[Mapping[str, Any]],
        key_fields: Sequence[str],
        create_if_missing: bool = True,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        """
        Insert rows when no match exists; update rows when key_fields match.
        """

        operation = "upsert_rows"
        provider = self._normalize_provider(provider)

        context_result = self._validate_task_context(context)
        if not context_result["success"]:
            return context_result

        if not key_fields:
            return self._error_result(
                message="key_fields are required for upsert.",
                code="missing_key_fields",
            )

        read_result = self.read_rows(
            context=context,
            provider=provider,
            target=target,
            limit=self.config.max_rows_per_read,
        )
        if not read_result["success"]:
            return read_result

        existing_rows = read_result["data"].get("rows", [])
        incoming_rows = self.normalize_rows(rows)

        existing_index: Dict[Tuple[str, ...], Dict[str, Any]] = {}
        for existing in existing_rows:
            existing_index[self._row_key(existing, key_fields)] = existing

        rows_to_insert: List[Dict[str, Any]] = []
        rows_to_update: List[Dict[str, Any]] = []

        for row in incoming_rows:
            key = self._row_key(row, key_fields)
            if key in existing_index:
                rows_to_update.append(row)
            else:
                rows_to_insert.append(row)

        if dry_run:
            return self._safe_result(
                message="Dry run completed. No rows were upserted.",
                data={
                    "provider": provider,
                    "insert_count": 0,
                    "update_count": 0,
                    "would_insert": len(rows_to_insert),
                    "would_update": len(rows_to_update),
                    "dry_run": True,
                },
                metadata={"operation": operation},
            )

        inserted_count = 0
        updated_count = 0

        if rows_to_insert:
            write_result = self.write_rows(
                context=context,
                provider=provider,
                target=target,
                rows=rows_to_insert,
                create_if_missing=create_if_missing,
                dry_run=False,
            )
            if not write_result["success"]:
                return write_result
            inserted_count = int(write_result["data"].get("written_count", 0))

        for row in rows_to_update:
            match = {field: row.get(field) for field in key_fields}
            updates = {key: value for key, value in row.items() if key not in key_fields}
            if updates:
                update_result = self.update_rows(
                    context=context,
                    provider=provider,
                    target=target,
                    match=match,
                    updates=updates,
                    dry_run=False,
                )
                if not update_result["success"]:
                    return update_result
                updated_count += int(update_result["data"].get("updated_count", 0))

        result = self._safe_result(
            message="Rows upserted successfully.",
            data={
                "provider": provider,
                "inserted_count": inserted_count,
                "updated_count": updated_count,
                "incoming_count": len(incoming_rows),
                "dry_run": False,
            },
            metadata={"operation": operation},
        )

        context_valid = self._validate_task_context(context)
        ctx = context_valid.get("data", {}) if context_valid["success"] else {}
        target_obj = self._build_target(provider, target).get("data", {}).get("target_obj")
        if target_obj:
            self._after_operation(ctx, operation, provider, target_obj, result)
        return result

    def find_duplicates(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: Mapping[str, Any],
        key_fields: Sequence[str],
        case_sensitive: bool = False,
        trim_whitespace: bool = True,
        normalize_phone: bool = True,
        filters: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Find duplicate rows by one or more key fields.
        """

        operation = "find_duplicates"
        provider = self._normalize_provider(provider)

        context_result = self._validate_task_context(context)
        if not context_result["success"]:
            return context_result
        ctx = context_result["data"]

        if not key_fields:
            return self._error_result(
                message="At least one key field is required for duplicate checks.",
                code="missing_key_fields",
            )

        read_result = self.read_rows(
            context=context,
            provider=provider,
            target=target,
            limit=self.config.max_rows_per_read,
            filters=filters,
        )
        if not read_result["success"]:
            return read_result

        rows = read_result["data"].get("rows", [])
        config = DuplicateCheckConfig(
            key_fields=list(key_fields),
            case_sensitive=case_sensitive,
            trim_whitespace=trim_whitespace,
            normalize_phone=normalize_phone,
        )

        duplicates = self._find_duplicates_in_rows(rows, config)

        target_obj = self._build_target(provider, target)["data"]["target_obj"]
        result = self._safe_result(
            message="Duplicate check completed.",
            data={
                "provider": provider,
                "row_count": len(rows),
                "duplicate_count": len(duplicates),
                "duplicates": duplicates,
                "key_fields": list(key_fields),
            },
            metadata={"operation": operation},
        )

        self._after_operation(ctx, operation, provider, target_obj, result)
        return result

    def export_rows(
        self,
        context: Mapping[str, Any],
        rows: Optional[Sequence[Mapping[str, Any]]] = None,
        provider: Optional[str] = None,
        target: Optional[Mapping[str, Any]] = None,
        export_format: str = "csv",
        file_name: Optional[str] = None,
        overwrite: bool = False,
        filters: Optional[Mapping[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Export rows to CSV, JSON, or XLSX.

        Rows can be provided directly, or loaded from provider + target.
        """

        operation = "export_rows"

        context_result = self._validate_task_context(context)
        if not context_result["success"]:
            return context_result
        ctx = context_result["data"]

        export_format = self._normalize_export_format(export_format)

        if self._requires_security_check(operation, provider or "local_export", ctx, target):
            approval = self._request_security_approval(
                context=ctx,
                operation=operation,
                provider=provider or "local_export",
                target=target,
                payload_summary={
                    "export_format": export_format,
                    "file_name": file_name,
                    "direct_rows": rows is not None,
                },
            )
            if not approval["success"]:
                return approval

        try:
            if rows is None:
                if not provider or not target:
                    return self._error_result(
                        message="Either rows or provider + target must be provided for export.",
                        code="missing_export_source",
                    )

                read_result = self.read_rows(
                    context=context,
                    provider=provider,
                    target=target,
                    filters=filters,
                )
                if not read_result["success"]:
                    return read_result
                export_rows = read_result["data"].get("rows", [])
                provider_name = self._normalize_provider(provider)
            else:
                export_rows = self.normalize_rows(rows)
                provider_name = provider or "direct_rows"

            export_path = self._write_export_file(
                context=ctx,
                rows=export_rows,
                export_format=export_format,
                file_name=file_name,
                overwrite=overwrite,
            )

            result = self._safe_result(
                message="Rows exported successfully.",
                data={
                    "provider": provider_name,
                    "row_count": len(export_rows),
                    "export_format": export_format,
                    "export_path": str(export_path),
                    "file_name": export_path.name,
                },
                metadata={"operation": operation},
            )

            target_obj = None
            if provider and target:
                target_obj_result = self._build_target(self._normalize_provider(provider), target)
                if target_obj_result["success"]:
                    target_obj = target_obj_result["data"]["target_obj"]

            self._after_operation(ctx, operation, provider_name, target_obj, result)
            return result

        except Exception as exc:
            result = self._error_result(
                message="Failed to export rows.",
                code="export_rows_failed",
                details={"exception": str(exc), "export_format": export_format},
                metadata={"operation": operation},
            )
            self._after_operation(ctx, operation, provider or "direct_rows", None, result)
            return result

    def normalize_rows(self, rows: Any) -> List[Dict[str, Any]]:
        """
        Normalize row values into JSON/spreadsheet-safe dictionaries.
        """

        normalized = _ensure_list_of_dicts(rows)
        clean_rows: List[Dict[str, Any]] = []

        for row in normalized:
            clean_row: Dict[str, Any] = {}
            for key, value in row.items():
                if key is None:
                    continue
                clean_key = str(key).strip()
                if not clean_key:
                    continue
                clean_row[clean_key] = _flatten_value(value)
            clean_rows.append(clean_row)

        return clean_rows

    # -------------------------------------------------------------------------
    # Google Sheets implementation
    # -------------------------------------------------------------------------

    def _get_google_client(self, target: SheetTarget) -> Any:
        if gspread is None:
            raise RuntimeError(
                "Google Sheets support requires optional dependencies: "
                "pip install gspread google-auth"
            )

        credentials_json = target.credentials_json

        if not credentials_json and target.service_account_env:
            credentials_json = _safe_read_json_env(target.service_account_env)

        if credentials_json:
            return gspread.service_account_from_dict(credentials_json)

        if target.credentials_file:
            return gspread.service_account(filename=target.credentials_file)

        default_credentials_file = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
        if default_credentials_file:
            return gspread.service_account(filename=default_credentials_file)

        raise RuntimeError(
            "Google Sheets credentials missing. Provide credentials_json, "
            "credentials_file, service_account_env, or GOOGLE_APPLICATION_CREDENTIALS."
        )

    def _get_google_worksheet(self, target: SheetTarget) -> Any:
        if not target.spreadsheet_id and not target.spreadsheet_name:
            raise ValueError("Google Sheets target requires spreadsheet_id or spreadsheet_name.")

        client = self._get_google_client(target)

        if target.spreadsheet_id:
            spreadsheet = client.open_by_key(target.spreadsheet_id)
        else:
            spreadsheet = client.open(target.spreadsheet_name)

        if target.worksheet_name:
            try:
                return spreadsheet.worksheet(target.worksheet_name)
            except Exception:
                if target.extra.get("create_worksheet_if_missing"):
                    return spreadsheet.add_worksheet(
                        title=target.worksheet_name,
                        rows=int(target.extra.get("initial_rows", 1000)),
                        cols=int(target.extra.get("initial_cols", 26)),
                    )
                raise

        return spreadsheet.sheet1

    def _google_read_rows(self, target: SheetTarget) -> List[Dict[str, Any]]:
        worksheet = self._get_google_worksheet(target)
        records = worksheet.get_all_records()
        return [dict(row) for row in records]

    def _google_append_rows(
        self,
        target: SheetTarget,
        rows: Sequence[Mapping[str, Any]],
        create_if_missing: bool = True,
    ) -> int:
        worksheet = self._get_google_worksheet(target)
        normalized_rows = self.normalize_rows(rows)
        headers = worksheet.row_values(1)

        if not headers:
            headers = self._collect_headers(normalized_rows)
            if headers:
                worksheet.append_row(headers)

        missing_headers = [header for header in self._collect_headers(normalized_rows) if header not in headers]
        if missing_headers:
            headers.extend(missing_headers)
            worksheet.update("1:1", [headers])

        values = [[row.get(header, "") for header in headers] for row in normalized_rows]
        if values:
            worksheet.append_rows(values, value_input_option="USER_ENTERED")
        return len(values)

    def _google_update_rows(
        self,
        target: SheetTarget,
        match: Mapping[str, Any],
        updates: Mapping[str, Any],
        dry_run: bool = False,
    ) -> int:
        worksheet = self._get_google_worksheet(target)
        all_values = worksheet.get_all_values()

        if not all_values:
            return 0

        headers = all_values[0]
        updated_count = 0
        rows_to_update: List[Tuple[int, List[Any]]] = []

        for row_number, values in enumerate(all_values[1:], start=2):
            row_dict = self._row_from_headers(headers, values)
            if self._matches(row_dict, match):
                updated_row = dict(row_dict)
                updated_row.update({str(k): _flatten_value(v) for k, v in updates.items()})

                for field in updated_row.keys():
                    if field not in headers:
                        headers.append(field)

                row_values = [updated_row.get(header, "") for header in headers]
                rows_to_update.append((row_number, row_values))
                updated_count += 1

        if dry_run:
            return updated_count

        if updated_count:
            worksheet.update("1:1", [headers])
            for row_number, row_values in rows_to_update:
                cell_range = f"A{row_number}:{self._column_letter(len(headers))}{row_number}"
                worksheet.update(cell_range, [row_values])

        return updated_count

    # -------------------------------------------------------------------------
    # Airtable implementation
    # -------------------------------------------------------------------------

    def _get_airtable_api_key(self, target: SheetTarget) -> str:
        api_key = target.api_key
        if not api_key and target.api_key_env:
            api_key = os.getenv(target.api_key_env)
        if not api_key:
            api_key = os.getenv("AIRTABLE_API_KEY")
        if not api_key:
            raise RuntimeError(
                "Airtable API key missing. Provide api_key, api_key_env, or AIRTABLE_API_KEY."
            )
        return api_key

    def _airtable_headers(self, target: SheetTarget) -> Dict[str, str]:
        api_key = self._get_airtable_api_key(target)
        return {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }

    def _airtable_url(self, target: SheetTarget) -> str:
        if not target.base_id:
            raise ValueError("Airtable target requires base_id.")
        if not target.table_name:
            raise ValueError("Airtable target requires table_name.")

        table = str(target.table_name).replace("/", "%2F")
        return f"{target.base_url.rstrip('/')}/{target.base_id}/{table}"

    def _airtable_read_rows(
        self,
        target: SheetTarget,
        limit: int,
        offset: int = 0,
    ) -> List[Dict[str, Any]]:
        if requests is None:
            raise RuntimeError("Airtable support requires optional dependency: pip install requests")

        url = self._airtable_url(target)
        headers = self._airtable_headers(target)

        params: Dict[str, Any] = {
            "pageSize": min(100, max(1, limit)),
        }
        if target.view:
            params["view"] = target.view

        rows: List[Dict[str, Any]] = []
        airtable_offset: Optional[str] = None

        while len(rows) < limit:
            if airtable_offset:
                params["offset"] = airtable_offset

            response = requests.get(
                url,
                headers=headers,
                params=params,
                timeout=self.config.default_timeout_seconds,
            )
            response.raise_for_status()
            payload = response.json()

            for record in payload.get("records", []):
                fields = dict(record.get("fields", {}))
                fields["_airtable_record_id"] = record.get("id")
                fields["_airtable_created_time"] = record.get("createdTime")
                rows.append(fields)
                if len(rows) >= limit:
                    break

            airtable_offset = payload.get("offset")
            if not airtable_offset:
                break

        if offset:
            rows = rows[offset:]

        return rows[:limit]

    def _airtable_create_rows(
        self,
        target: SheetTarget,
        rows: Sequence[Mapping[str, Any]],
    ) -> int:
        if requests is None:
            raise RuntimeError("Airtable support requires optional dependency: pip install requests")

        url = self._airtable_url(target)
        headers = self._airtable_headers(target)
        normalized_rows = self.normalize_rows(rows)

        written = 0
        for chunk in self._chunk_list(normalized_rows, 10):
            payload = {
                "records": [
                    {"fields": self._strip_internal_fields(row)}
                    for row in chunk
                ]
            }

            response = requests.post(
                url,
                headers=headers,
                json=payload,
                timeout=self.config.default_timeout_seconds,
            )
            response.raise_for_status()
            created = response.json().get("records", [])
            written += len(created)

        return written

    def _airtable_update_rows(
        self,
        target: SheetTarget,
        match: Mapping[str, Any],
        updates: Mapping[str, Any],
        dry_run: bool = False,
    ) -> int:
        if requests is None:
            raise RuntimeError("Airtable support requires optional dependency: pip install requests")

        existing = self._airtable_read_rows(target, limit=self.config.max_rows_per_read)
        matched_records = [
            row for row in existing
            if self._matches(row, match) and row.get("_airtable_record_id")
        ]

        if dry_run:
            return len(matched_records)

        url = self._airtable_url(target)
        headers = self._airtable_headers(target)
        clean_updates = self._strip_internal_fields(
            {str(k): _flatten_value(v) for k, v in updates.items()}
        )

        updated_count = 0
        for chunk in self._chunk_list(matched_records, 10):
            payload = {
                "records": [
                    {
                        "id": row["_airtable_record_id"],
                        "fields": clean_updates,
                    }
                    for row in chunk
                ]
            }

            response = requests.patch(
                url,
                headers=headers,
                json=payload,
                timeout=self.config.default_timeout_seconds,
            )
            response.raise_for_status()
            updated_count += len(response.json().get("records", []))

        return updated_count

    # -------------------------------------------------------------------------
    # Excel implementation
    # -------------------------------------------------------------------------

    def _excel_read_rows(self, target: SheetTarget) -> List[Dict[str, Any]]:
        if load_workbook is None:
            raise RuntimeError("Excel support requires optional dependency: pip install openpyxl")

        file_path = self._validate_local_file_target(target, extension=".xlsx", must_exist=True)
        workbook = load_workbook(filename=str(file_path), data_only=True)
        sheet_name = target.sheet_name or target.worksheet_name or workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows: List[Dict[str, Any]] = []

        for values in rows[1:]:
            row = self._row_from_headers(headers, values)
            data_rows.append(row)

        return data_rows

    def _excel_append_rows(
        self,
        target: SheetTarget,
        rows: Sequence[Mapping[str, Any]],
        create_if_missing: bool = True,
    ) -> int:
        if Workbook is None or load_workbook is None:
            raise RuntimeError("Excel support requires optional dependency: pip install openpyxl")

        file_path = self._validate_local_file_target(
            target,
            extension=".xlsx",
            must_exist=not create_if_missing,
        )

        normalized_rows = self.normalize_rows(rows)

        if file_path.exists():
            workbook = load_workbook(filename=str(file_path))
        else:
            workbook = Workbook()

        sheet_name = target.sheet_name or target.worksheet_name or "Sheet1"

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        if workbook.active and workbook.active.title == "Sheet" and sheet_name != "Sheet":
            default_sheet = workbook["Sheet"] if "Sheet" in workbook.sheetnames else None
            if default_sheet is not None and default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
                try:
                    workbook.remove(default_sheet)
                except Exception:
                    pass

        existing_headers = self._excel_headers(sheet)
        incoming_headers = self._collect_headers(normalized_rows)

        if not existing_headers:
            existing_headers = incoming_headers
            sheet.append(existing_headers)
        else:
            missing_headers = [header for header in incoming_headers if header not in existing_headers]
            if missing_headers:
                existing_headers.extend(missing_headers)
                for col_index, header in enumerate(existing_headers, start=1):
                    sheet.cell(row=1, column=col_index, value=header)

        for row in normalized_rows:
            sheet.append([row.get(header, "") for header in existing_headers])

        _safe_mkdir(file_path.parent)
        workbook.save(str(file_path))
        return len(normalized_rows)

    def _excel_update_rows(
        self,
        target: SheetTarget,
        match: Mapping[str, Any],
        updates: Mapping[str, Any],
        dry_run: bool = False,
    ) -> int:
        if load_workbook is None:
            raise RuntimeError("Excel support requires optional dependency: pip install openpyxl")

        file_path = self._validate_local_file_target(target, extension=".xlsx", must_exist=True)
        workbook = load_workbook(filename=str(file_path))
        sheet_name = target.sheet_name or target.worksheet_name or workbook.sheetnames[0]

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        headers = self._excel_headers(sheet)

        if not headers:
            return 0

        for field in updates.keys():
            clean_field = str(field)
            if clean_field not in headers:
                headers.append(clean_field)
                sheet.cell(row=1, column=len(headers), value=clean_field)

        updated_count = 0

        for row_index in range(2, sheet.max_row + 1):
            row_dict = {}
            for col_index, header in enumerate(headers, start=1):
                row_dict[header] = sheet.cell(row=row_index, column=col_index).value

            if self._matches(row_dict, match):
                updated_count += 1
                if not dry_run:
                    for field, value in updates.items():
                        col_index = headers.index(str(field)) + 1
                        sheet.cell(row=row_index, column=col_index, value=_flatten_value(value))

        if not dry_run:
            workbook.save(str(file_path))

        return updated_count

    def _excel_headers(self, sheet: Any) -> List[str]:
        headers: List[str] = []
        if sheet.max_row < 1:
            return headers

        for col_index in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col_index).value
            if value is not None and str(value).strip():
                headers.append(str(value).strip())

        return headers

    # -------------------------------------------------------------------------
    # CSV implementation
    # -------------------------------------------------------------------------

    def _csv_read_rows(self, target: SheetTarget) -> List[Dict[str, Any]]:
        file_path = self._validate_local_file_target(target, extension=".csv", must_exist=True)

        with file_path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            return [dict(row) for row in reader]

    def _csv_append_rows(
        self,
        target: SheetTarget,
        rows: Sequence[Mapping[str, Any]],
        create_if_missing: bool = True,
    ) -> int:
        file_path = self._validate_local_file_target(
            target,
            extension=".csv",
            must_exist=not create_if_missing,
        )

        normalized_rows = self.normalize_rows(rows)
        existing_rows: List[Dict[str, Any]] = []

        if file_path.exists():
            existing_rows = self._csv_read_rows(target)

        headers = self._collect_headers(existing_rows + normalized_rows)
        _safe_mkdir(file_path.parent)

        file_exists = file_path.exists()
        should_rewrite = False

        if file_exists:
            existing_headers = []
            with file_path.open("r", newline="", encoding="utf-8-sig") as handle:
                reader = csv.reader(handle)
                try:
                    existing_headers = next(reader)
                except StopIteration:
                    existing_headers = []

            if set(headers) != set(existing_headers) or headers != existing_headers:
                should_rewrite = True

        if should_rewrite:
            all_rows = existing_rows + normalized_rows
            self._write_csv(file_path, all_rows, headers)
        else:
            mode = "a" if file_exists else "w"
            with file_path.open(mode, newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                if not file_exists or file_path.stat().st_size == 0:
                    writer.writeheader()
                for row in normalized_rows:
                    writer.writerow({header: row.get(header, "") for header in headers})

        return len(normalized_rows)

    def _csv_update_rows(
        self,
        target: SheetTarget,
        match: Mapping[str, Any],
        updates: Mapping[str, Any],
        dry_run: bool = False,
    ) -> int:
        file_path = self._validate_local_file_target(target, extension=".csv", must_exist=True)
        rows = self._csv_read_rows(target)
        headers = self._collect_headers(rows + [dict(updates)])

        updated_count = 0
        updated_rows: List[Dict[str, Any]] = []

        for row in rows:
            clean_row = dict(row)
            if self._matches(clean_row, match):
                updated_count += 1
                if not dry_run:
                    clean_row.update({str(k): _flatten_value(v) for k, v in updates.items()})
            updated_rows.append(clean_row)

        if not dry_run:
            self._write_csv(file_path, updated_rows, headers)

        return updated_count

    # -------------------------------------------------------------------------
    # Duplicate logic
    # -------------------------------------------------------------------------

    def _dedupe_before_write(
        self,
        context: Mapping[str, Any],
        provider: str,
        target: SheetTarget,
        incoming_rows: List[Dict[str, Any]],
        dedupe: Mapping[str, Any],
    ) -> Dict[str, Any]:
        key_fields = dedupe.get("key_fields") or dedupe.get("fields")
        if not key_fields:
            return self._error_result(
                message="Dedupe config requires key_fields.",
                code="missing_dedupe_key_fields",
            )

        if isinstance(key_fields, str):
            key_fields = [key_fields]

        existing_result = self.read_rows(
            context=context,
            provider=provider,
            target=asdict(target),
            limit=self.config.max_rows_per_read,
        )
        if not existing_result["success"]:
            return existing_result

        existing_rows = existing_result["data"].get("rows", [])
        config = DuplicateCheckConfig(
            key_fields=list(key_fields),
            case_sensitive=bool(dedupe.get("case_sensitive", False)),
            trim_whitespace=bool(dedupe.get("trim_whitespace", True)),
            normalize_phone=bool(dedupe.get("normalize_phone", True)),
            ignore_empty_keys=bool(dedupe.get("ignore_empty_keys", True)),
        )

        existing_keys = {
            self._duplicate_key(row, config)
            for row in existing_rows
            if self._duplicate_key(row, config)
        }

        seen_incoming = set()
        rows_to_write: List[Dict[str, Any]] = []
        skipped_duplicates: List[Dict[str, Any]] = []

        for index, row in enumerate(incoming_rows):
            key = self._duplicate_key(row, config)
            if not key and config.ignore_empty_keys:
                rows_to_write.append(row)
                continue

            if key in existing_keys or key in seen_incoming:
                skipped_duplicates.append(
                    {
                        "incoming_index": index,
                        "key": key,
                        "row": row,
                    }
                )
                continue

            seen_incoming.add(key)
            rows_to_write.append(row)

        mode = str(dedupe.get("mode", "skip")).lower()
        if mode == "block" and skipped_duplicates:
            return self._error_result(
                message="Duplicate rows found. Write blocked.",
                code="duplicates_found",
                details={
                    "duplicate_count": len(skipped_duplicates),
                    "duplicates": skipped_duplicates,
                    "key_fields": list(key_fields),
                },
            )

        return self._safe_result(
            message="Dedupe check completed.",
            data={
                "rows_to_write": rows_to_write,
                "skipped_duplicates": skipped_duplicates,
                "duplicate_count": len(skipped_duplicates),
            },
        )

    def _find_duplicates_in_rows(
        self,
        rows: Sequence[Mapping[str, Any]],
        config: DuplicateCheckConfig,
    ) -> List[Dict[str, Any]]:
        seen: Dict[Tuple[str, ...], List[int]] = {}
        row_map: Dict[Tuple[str, ...], List[Dict[str, Any]]] = {}

        for index, row in enumerate(rows):
            key = self._duplicate_key(row, config)
            if not key and config.ignore_empty_keys:
                continue

            seen.setdefault(key, []).append(index)
            row_map.setdefault(key, []).append(dict(row))

        duplicates: List[Dict[str, Any]] = []
        for key, indexes in seen.items():
            if len(indexes) > 1:
                duplicates.append(
                    {
                        "key": key,
                        "indexes": indexes,
                        "count": len(indexes),
                        "rows": row_map.get(key, []),
                    }
                )

        return duplicates

    def _duplicate_key(
        self,
        row: Mapping[str, Any],
        config: DuplicateCheckConfig,
    ) -> Tuple[str, ...]:
        key_parts: List[str] = []

        for field in config.key_fields:
            value = row.get(field)

            if value is None:
                value = ""

            if config.normalize_phone and self._looks_like_phone_field(field):
                value = _normalize_phone(value)
            else:
                value = str(value)

            if config.trim_whitespace:
                value = str(value).strip()

            if not config.case_sensitive:
                value = str(value).lower()

            key_parts.append(str(value))

        if config.ignore_empty_keys and not any(key_parts):
            return tuple()

        return tuple(key_parts)

    # -------------------------------------------------------------------------
    # Export implementation
    # -------------------------------------------------------------------------

    def _write_export_file(
        self,
        context: Mapping[str, Any],
        rows: Sequence[Mapping[str, Any]],
        export_format: str,
        file_name: Optional[str],
        overwrite: bool,
    ) -> Path:
        export_format = self._normalize_export_format(export_format)
        normalized_rows = self.normalize_rows(rows)

        base_dir = Path(self.config.export_base_dir).resolve()
        user_id = self._safe_slug(context.get("user_id", "unknown_user"))
        workspace_id = self._safe_slug(context.get("workspace_id", "unknown_workspace"))
        export_dir = base_dir / workspace_id / user_id
        _safe_mkdir(export_dir)

        if file_name:
            safe_name = self._safe_export_filename(file_name, export_format)
        else:
            safe_name = f"sheet_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.{export_format}"

        export_path = export_dir / safe_name

        if export_path.exists() and not overwrite:
            stem = export_path.stem
            suffix = export_path.suffix
            export_path = export_dir / f"{stem}_{uuid.uuid4().hex[:8]}{suffix}"

        if export_format == "csv":
            headers = self._collect_headers(normalized_rows)
            self._write_csv(export_path, normalized_rows, headers)
        elif export_format == "json":
            with export_path.open("w", encoding="utf-8") as handle:
                json.dump(
                    {
                        "metadata": {
                            "exported_at": _utc_now_iso(),
                            "row_count": len(normalized_rows),
                            "user_id": context.get("user_id"),
                            "workspace_id": context.get("workspace_id"),
                        },
                        "rows": normalized_rows,
                    },
                    handle,
                    ensure_ascii=False,
                    indent=2,
                    default=str,
                )
        elif export_format == "xlsx":
            if Workbook is None:
                raise RuntimeError("XLSX export requires optional dependency: pip install openpyxl")

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Export"
            headers = self._collect_headers(normalized_rows)
            if headers:
                sheet.append(headers)
                for row in normalized_rows:
                    sheet.append([row.get(header, "") for header in headers])
            workbook.save(str(export_path))
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

        return export_path

    # -------------------------------------------------------------------------
    # Common helpers
    # -------------------------------------------------------------------------

    def _normalize_provider(self, provider: str) -> str:
        normalized = str(provider or "").strip().lower().replace("-", "_").replace(" ", "_")

        aliases = {
            "google": "google_sheets",
            "gsheets": "google_sheets",
            "gspread": "google_sheets",
            "sheet": "google_sheets",
            "sheets": "google_sheets",
            "air_table": "airtable",
            "xlsx": "excel",
            "xls": "excel",
        }

        normalized = aliases.get(normalized, normalized)

        if normalized not in SUPPORTED_PROVIDERS:
            raise ValueError(f"Unsupported provider '{provider}'. Supported: {sorted(SUPPORTED_PROVIDERS)}")

        return normalized

    def _normalize_export_format(self, export_format: str) -> str:
        normalized = str(export_format or "csv").strip().lower().replace(".", "")
        aliases = {
            "excel": "xlsx",
            "xls": "xlsx",
        }
        normalized = aliases.get(normalized, normalized)

        if normalized not in {"csv", "json", "xlsx"}:
            raise ValueError("export_format must be csv, json, or xlsx.")

        return normalized

    def _build_target(self, provider: str, target: Mapping[str, Any]) -> Dict[str, Any]:
        if not isinstance(target, Mapping):
            return self._error_result(
                message="target must be a dictionary.",
                code="invalid_target",
            )

        try:
            target_data = dict(target)
            target_data["provider"] = provider

            known_fields = {
                "provider",
                "spreadsheet_id",
                "spreadsheet_name",
                "worksheet_name",
                "base_id",
                "table_name",
                "view",
                "base_url",
                "file_path",
                "sheet_name",
                "credentials_json",
                "credentials_file",
                "service_account_env",
                "api_key",
                "api_key_env",
                "extra",
            }

            extra = dict(target_data.get("extra") or {})
            for key in list(target_data.keys()):
                if key not in known_fields:
                    extra[key] = target_data.pop(key)

            target_data["extra"] = extra

            target_obj = SheetTarget(**target_data)
            self._validate_target_by_provider(target_obj)
            return self._safe_result(
                message="Target validated.",
                data={"target_obj": target_obj},
            )
        except Exception as exc:
            return self._error_result(
                message="Invalid sheet target.",
                code="invalid_target",
                details={"exception": str(exc), "provider": provider},
            )

    def _validate_target_by_provider(self, target: SheetTarget) -> None:
        provider = target.provider

        if provider == "google_sheets":
            if not target.spreadsheet_id and not target.spreadsheet_name:
                raise ValueError("Google Sheets target requires spreadsheet_id or spreadsheet_name.")

        elif provider == "airtable":
            if not target.base_id:
                raise ValueError("Airtable target requires base_id.")
            if not target.table_name:
                raise ValueError("Airtable target requires table_name.")

        elif provider in {"excel", "csv"}:
            if not self.config.allow_local_file_access:
                raise PermissionError("Local file access is disabled for SheetConnector.")
            if not target.file_path:
                raise ValueError(f"{provider} target requires file_path.")

        else:
            raise ValueError(f"Unsupported provider: {provider}")

    def _validate_local_file_target(
        self,
        target: SheetTarget,
        extension: str,
        must_exist: bool,
    ) -> Path:
        if not self.config.allow_local_file_access:
            raise PermissionError("Local file access is disabled.")

        if not target.file_path:
            raise ValueError("file_path is required.")

        file_path = _safe_file_path(target.file_path)

        if file_path.suffix.lower() != extension:
            raise ValueError(f"Expected {extension} file, got {file_path.suffix}")

        if file_path.suffix.lower() not in self.config.allowed_local_extensions:
            raise ValueError(f"File extension not allowed: {file_path.suffix}")

        if must_exist and not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        return file_path

    def _safe_limit(self, limit: Optional[int]) -> int:
        if limit is None:
            return self.config.max_rows_per_read

        try:
            value = int(limit)
        except Exception:
            value = self.config.max_rows_per_read

        return max(1, min(value, self.config.max_rows_per_read))

    def _apply_filters(
        self,
        rows: Sequence[Mapping[str, Any]],
        filters: Mapping[str, Any],
    ) -> List[Dict[str, Any]]:
        if not filters:
            return [dict(row) for row in rows]

        filtered: List[Dict[str, Any]] = []

        for row in rows:
            if self._matches(row, filters):
                filtered.append(dict(row))

        return filtered

    def _matches(self, row: Mapping[str, Any], criteria: Mapping[str, Any]) -> bool:
        for key, expected in criteria.items():
            actual = row.get(key)

            if isinstance(expected, Mapping):
                if not self._matches_operator(actual, expected):
                    return False
            elif isinstance(expected, (list, tuple, set)):
                if actual not in expected and str(actual) not in {str(item) for item in expected}:
                    return False
            else:
                if str(actual).strip() != str(expected).strip():
                    return False

        return True

    def _matches_operator(self, actual: Any, expected: Mapping[str, Any]) -> bool:
        for operator, value in expected.items():
            op = str(operator).lower()

            if op in {"eq", "equals"}:
                if str(actual).strip() != str(value).strip():
                    return False

            elif op in {"ne", "not_equals"}:
                if str(actual).strip() == str(value).strip():
                    return False

            elif op in {"contains"}:
                if str(value).lower() not in str(actual).lower():
                    return False

            elif op in {"starts_with"}:
                if not str(actual).lower().startswith(str(value).lower()):
                    return False

            elif op in {"ends_with"}:
                if not str(actual).lower().endswith(str(value).lower()):
                    return False

            elif op in {"empty", "is_empty"}:
                is_empty = actual in (None, "")
                if bool(value) != is_empty:
                    return False

            elif op in {"gt", "gte", "lt", "lte"}:
                try:
                    actual_num = float(actual)
                    expected_num = float(value)
                except Exception:
                    return False

                if op == "gt" and not actual_num > expected_num:
                    return False
                if op == "gte" and not actual_num >= expected_num:
                    return False
                if op == "lt" and not actual_num < expected_num:
                    return False
                if op == "lte" and not actual_num <= expected_num:
                    return False

            else:
                return False

        return True

    def _row_from_headers(
        self,
        headers: Sequence[Any],
        values: Sequence[Any],
    ) -> Dict[str, Any]:
        row: Dict[str, Any] = {}
        for index, header in enumerate(headers):
            clean_header = str(header).strip() if header is not None else ""
            if not clean_header:
                continue
            row[clean_header] = values[index] if index < len(values) else ""
        return row

    def _collect_headers(self, rows: Sequence[Mapping[str, Any]]) -> List[str]:
        headers: List[str] = []
        seen = set()

        for row in rows:
            for key in row.keys():
                clean_key = str(key).strip()
                if not clean_key or clean_key in seen:
                    continue
                headers.append(clean_key)
                seen.add(clean_key)

        return headers

    def _write_csv(
        self,
        file_path: Path,
        rows: Sequence[Mapping[str, Any]],
        headers: Sequence[str],
    ) -> None:
        _safe_mkdir(file_path.parent)
        with file_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(headers))
            writer.writeheader()
            for row in rows:
                writer.writerow({header: row.get(header, "") for header in headers})

    def _strip_internal_fields(self, row: Mapping[str, Any]) -> Dict[str, Any]:
        return {
            key: value
            for key, value in row.items()
            if not str(key).startswith("_airtable_")
        }

    def _row_key(self, row: Mapping[str, Any], key_fields: Sequence[str]) -> Tuple[str, ...]:
        return tuple(str(row.get(field, "")).strip().lower() for field in key_fields)

    def _chunk_list(self, items: Sequence[Dict[str, Any]], chunk_size: int) -> Iterable[List[Dict[str, Any]]]:
        for index in range(0, len(items), chunk_size):
            yield list(items[index:index + chunk_size])

    def _looks_like_phone_field(self, field: str) -> bool:
        field_lower = str(field).lower()
        return any(token in field_lower for token in ["phone", "mobile", "whatsapp", "number", "contact"])

    def _column_letter(self, column_number: int) -> str:
        result = ""
        while column_number:
            column_number, remainder = divmod(column_number - 1, 26)
            result = chr(65 + remainder) + result
        return result or "A"

    def _safe_slug(self, value: Any) -> str:
        text = str(value or "unknown").strip()
        text = re.sub(r"[^a-zA-Z0-9._-]+", "_", text)
        return text[:120] or "unknown"

    def _safe_export_filename(self, file_name: str, export_format: str) -> str:
        name = Path(file_name).name
        name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name)

        if not name:
            name = f"sheet_export_{uuid.uuid4().hex[:8]}.{export_format}"

        suffix = Path(name).suffix.lower().replace(".", "")
        if suffix != export_format:
            name = f"{Path(name).stem}.{export_format}"

        return name

    def _sanitize_target_for_logs(self, target: Mapping[str, Any]) -> Dict[str, Any]:
        sanitized = dict(target)

        sensitive_keys = {
            "api_key",
            "credentials_json",
            "credentials_file",
            "service_account_env",
            "api_key_env",
        }

        for key in list(sanitized.keys()):
            if key in sensitive_keys:
                sanitized[key] = _redact(sanitized[key])

        if isinstance(sanitized.get("extra"), Mapping):
            extra = dict(sanitized["extra"])
            for key in list(extra.keys()):
                if "key" in key.lower() or "secret" in key.lower() or "token" in key.lower():
                    extra[key] = _redact(extra[key])
            sanitized["extra"] = extra

        return _dict_without_none(sanitized)

    def _summarize_result_data(self, data: Any) -> Dict[str, Any]:
        if not isinstance(data, Mapping):
            return {}

        summary_keys = [
            "provider",
            "row_count",
            "written_count",
            "updated_count",
            "matched_count",
            "duplicate_count",
            "inserted_count",
            "export_format",
            "export_path",
            "file_name",
            "dry_run",
        ]

        return {
            key: data.get(key)
            for key in summary_keys
            if key in data
        }

    def _audit_start(
        self,
        context: Mapping[str, Any],
        operation: str,
        provider: str,
        target: Optional[SheetTarget],
        payload_summary: Optional[Mapping[str, Any]] = None,
    ) -> None:
        event = {
            "event_type": "sheet_operation_started",
            "operation": operation,
            "provider": provider,
            "context": {
                "user_id": context.get("user_id"),
                "workspace_id": context.get("workspace_id"),
                "request_id": context.get("request_id"),
                "actor_id": context.get("actor_id"),
            },
            "target": self._sanitize_target_for_logs(asdict(target) if target else {}),
            "payload_summary": dict(payload_summary or {}),
        }
        self._log_audit_event(event)
        self._emit_agent_event(event)

    def _after_operation(
        self,
        context: Mapping[str, Any],
        operation: str,
        provider: str,
        target: Optional[SheetTarget],
        result: Mapping[str, Any],
    ) -> None:
        event = {
            "event_type": "sheet_operation_completed",
            "operation": operation,
            "provider": provider,
            "success": bool(result.get("success")),
            "message": result.get("message"),
            "context": {
                "user_id": context.get("user_id"),
                "workspace_id": context.get("workspace_id"),
                "request_id": context.get("request_id"),
                "actor_id": context.get("actor_id"),
            },
            "target": self._sanitize_target_for_logs(asdict(target) if target else {}),
            "result_summary": self._summarize_result_data(result.get("data", {})),
            "error": result.get("error"),
        }

        self._log_audit_event(event)
        self._emit_agent_event(event)

        verification_payload = self._prepare_verification_payload(
            context=context,
            operation=operation,
            provider=provider,
            result=result,
            target=asdict(target) if target else None,
        )

        memory_payload = self._prepare_memory_payload(
            context=context,
            operation=operation,
            provider=provider,
            result=result,
            target=asdict(target) if target else None,
        )

        self._send_optional_payload(self.verification_client, verification_payload, "verification")
        self._send_optional_payload(self.memory_client, memory_payload, "memory")

    def _send_optional_payload(
        self,
        client: Optional[Any],
        payload: Mapping[str, Any],
        client_name: str,
    ) -> None:
        if client is None:
            return

        try:
            if hasattr(client, "record"):
                client.record(dict(payload))
            elif hasattr(client, "store"):
                client.store(dict(payload))
            elif hasattr(client, "submit"):
                client.submit(dict(payload))
            elif callable(client):
                client(dict(payload))
        except Exception as exc:
            self.logger.warning("Failed to send %s payload: %s", client_name, exc)


# =============================================================================
# Convenience factory for Agent Loader / Registry
# =============================================================================

def create_sheet_connector(
    config: Optional[Union[SheetConnectorConfig, Dict[str, Any]]] = None,
    **kwargs: Any,
) -> SheetConnector:
    """
    Factory used by future Agent Loader or Registry.
    """

    return SheetConnector(config=config, **kwargs)


def get_agent_manifest() -> Dict[str, Any]:
    """
    Static manifest for Master Agent / Agent Registry discovery.
    """

    return {
        "agent": "SheetConnector",
        "module": "workflow_agent",
        "file_path": "agents/workflow_agent/sheet_connector.py",
        "class_name": "SheetConnector",
        "factory": "create_sheet_connector",
        "purpose": "Reads/writes Google Sheets/Airtable/Excel rows, duplicate checks, exports.",
        "providers": sorted(SUPPORTED_PROVIDERS),
        "public_methods": [
            "health_check",
            "get_capabilities",
            "read_rows",
            "write_row",
            "write_rows",
            "update_rows",
            "upsert_rows",
            "find_duplicates",
            "export_rows",
            "normalize_rows",
        ],
        "requires_context": ["user_id", "workspace_id"],
        "security_hooks": [
            "_validate_task_context",
            "_requires_security_check",
            "_request_security_approval",
            "_prepare_verification_payload",
            "_prepare_memory_payload",
            "_emit_agent_event",
            "_log_audit_event",
            "_safe_result",
            "_error_result",
        ],
        "completion": {
            "agent_module": "Workflow Agent",
            "file_completed": "sheet_connector.py",
            "completion_percent": "47.6%",
            "next_recommended_file": "agents/workflow_agent/whatsapp_connector.py",
        },
    }


__all__ = [
    "SheetConnector",
    "SheetConnectorConfig",
    "SheetTarget",
    "DuplicateCheckConfig",
    "ExportConfig",
    "create_sheet_connector",
    "get_agent_manifest",
]


# =============================================================================
# Completion tracking
# =============================================================================
#
# Agent/Module: Workflow Agent
# File Completed: sheet_connector.py
# Completion: 47.6%
# Completed Files: ['workflow_agent.py', 'n8n_connector.py', 'workflow_builder.py', 'trigger_engine.py', 'action_router.py', 'app_connector.py', 'webhook_manager.py', 'form_pipeline.py', 'crm_connector.py', 'sheet_connector.py']
# Remaining Files: ['whatsapp_connector.py', 'email_connector.py', 'notification_engine.py', 'condition_engine.py', 'scheduler.py', 'workflow_monitor.py', 'retry_handler.py', 'workflow_templates.py', 'workflow_memory.py', 'approval_gate.py', 'config.py']
# Next Recommended File: agents/workflow_agent/whatsapp_connector.py
# FILE COMPLETE